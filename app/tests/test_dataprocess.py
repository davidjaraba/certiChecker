import os
import unittest
from unittest.mock import patch

import pytesseract
from openpyxl.reader.excel import load_workbook
from app.data_process import basic_process

url = 'https://raw.si.com'
origin_url = 'www.acciona.com'


class TestBasicProcess(unittest.TestCase):
    def setUp(self):
        # Print current working directory
        project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        print(f"Project root directory: {project_root}")

        # Construct the path to the Excel file relative to the project root
        file_path = os.path.join(project_root, 'tests/Glosario_Certificaciones.xlsx')
        print(f"Excel file path: {file_path}")

        certs = []
        wb = load_workbook(file_path)
        sheet = wb.active
        # Leer todas las filas de la columna B y convertirlas en un array de certificados
        for row in sheet.iter_rows(min_row=2, min_col=2, max_col=2, values_only=True):
            cell_value = row[0]
            if cell_value is not None:
                certs.append(cell_value)

        self.current_certs = certs

    @patch('app.data_process.save_founded_certs')
    def test_basic_process_txt(self, mock_save_founded_certs):

        res_certs = ['ISO 14001', 'ISO 9001']
        data_type = 'txt'
        txt_for_test = 'site_text.txt'
        needed_certs = []

        project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__))) + '/tests'
        txt_for_test = os.path.join(project_root, txt_for_test)

        mock_save_founded_certs.return_value = []

        try:
            needed_certs = basic_process(txt_for_test, data_type, url, origin_url, self.current_certs)
        except Exception as e:
            print(e)

        # Comparar el resultado con lo esperado
        self.assertCountEqual(res_certs, needed_certs)

    @patch('app.data_process.save_founded_certs')
    def test_basic_process_img_1(self, mock_save_founded_certs):
        res_certs = ['ISO 9001']
        data_type = 'img'
        img_for_test = 'iso9001.jpg'
        needed_certs = []

        project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__))) + '/tests'
        img_for_test = os.path.join(project_root, img_for_test)

        mock_save_founded_certs.return_value = []

        # Llamar a la función a probar
        try:
            needed_certs = basic_process(img_for_test, data_type, url, origin_url, self.current_certs)
        except Exception as e:
            print(e)

        # Comparar el resultado con lo esperado
        self.assertEqual(res_certs, needed_certs)

    @patch('app.data_process.save_founded_certs')
    def test_basic_process_img_2(self, mock_save_founded_certs):
        res_certs = ['Sustainalytics']
        data_type = 'img'
        img_for_test = 'sustainalytics.png'
        needed_certs = []

        project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__))) + '/tests'
        img_for_test = os.path.join(project_root, img_for_test)

        mock_save_founded_certs.return_value = []

        # Llamar a la función a probar
        try:
            needed_certs = basic_process(img_for_test, data_type, url, origin_url, self.current_certs)
        except Exception as e:
            print(e)

        # Comparar el resultado con lo esperado
        self.assertEqual(res_certs, needed_certs)

    # @patch('app.data_process.save_founded_certs')
    # def test_basic_process_img_3(self, mock_save_founded_certs):
    #     res_certs = ['UN GLOBAL COMPACT']
    #     data_type = 'img'
    #     img_for_test = 'globalcompact.jpg'
    #
    #     mock_save_founded_certs.return_value = []
    #
    #     # Llamar a la función a probar
    #     try:
    #         needed_certs = basic_process(img_for_test, data_type, url, origin_url, self.current_certs)
    #     except Exception as e:
    #         print(e)
    #
    #     # Comparar el resultado con lo esperado
    #     self.assertEqual(res_certs, needed_certs)

    @patch('app.data_process.save_founded_certs')
    def test_basic_process_img_4(self, mock_save_founded_certs):
        res_certs = ['Euronext']
        data_type = 'img'
        img_for_test = 'euronext.jpg'
        needed_certs = []

        project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__))) + '/tests'
        img_for_test = os.path.join(project_root, img_for_test)

        mock_save_founded_certs.return_value = []

        # Llamar a la función a probar
        try:
            needed_certs = basic_process(img_for_test, data_type, url, origin_url, self.current_certs)
        except Exception as e:
            print(e)

        # Comparar el resultado con lo esperado
        self.assertEqual(res_certs, needed_certs)

    @patch('app.data_process.save_founded_certs')
    def test_basic_process_img_5(self, mock_save_founded_certs):
        # res_certs = ['Dow Jones Sustainability Index']
        res_certs = ['S&P Global', 'Dow Jones Sustainability Index']
        data_type = 'img'
        img_for_test = 'dowjones.jpg'
        needed_certs = []

        project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__))) + '/tests'
        img_for_test = os.path.join(project_root, img_for_test)

        mock_save_founded_certs.return_value = []

        # Llamar a la función a probar
        try:
            needed_certs = basic_process(img_for_test, data_type, url, origin_url, self.current_certs)
        except Exception as e:
            print(e)

        # Comparar el resultado con lo esperado
        self.assertCountEqual(res_certs, needed_certs)

    @patch('app.data_process.save_founded_certs')
    def test_basic_process_img_6(self, mock_save_founded_certs):
        res_certs = ['Gaïa Rating']
        data_type = 'img'
        img_for_test = 'gaia.jpg'
        needed_certs = []

        project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__))) + '/tests'
        img_for_test = os.path.join(project_root, img_for_test)

        mock_save_founded_certs.return_value = []

        # Llamar a la función a probar
        try:
            needed_certs = basic_process(img_for_test, data_type, url, origin_url, self.current_certs)
        except Exception as e:
            print(e)

        # Comparar el resultado con lo esperado
        self.assertEqual(res_certs, needed_certs)

    @patch('app.data_process.save_founded_certs')
    def test_basic_process_img_7(self, mock_save_founded_certs):
        res_certs = ['Ecoact']
        data_type = 'img'
        img_for_test = 'ecoact.jpg'
        needed_certs = []

        project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__))) + '/tests'
        img_for_test = os.path.join(project_root, img_for_test)

        mock_save_founded_certs.return_value = []

        # Llamar a la función a probar
        try:
            needed_certs = basic_process(img_for_test, data_type, url, origin_url, self.current_certs)
        except Exception as e:
            print(e)

        # Comparar el resultado con lo esperado
        self.assertEqual(res_certs, needed_certs)

    @patch('app.data_process.save_founded_certs')
    def test_basic_process_img_8(self, mock_save_founded_certs):
        res_certs = ['S&P Global']
        data_type = 'img'
        img_for_test = 'sp.jpg'
        needed_certs = []

        project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__))) + '/tests'
        img_for_test = os.path.join(project_root, img_for_test)

        mock_save_founded_certs.return_value = []

        # Llamar a la función a probar
        try:
            needed_certs = basic_process(img_for_test, data_type, url, origin_url, self.current_certs)
        except Exception as e:
            print(e)

        # Comparar el resultado con lo esperado
        self.assertEqual(res_certs, needed_certs)

    @patch('app.data_process.save_founded_certs')
    def test_basic_process_img_9(self, mock_save_founded_certs):
        res_certs = ['Seal Awards']
        data_type = 'img'
        img_for_test = 'seal.jpg'
        needed_certs = []

        project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__))) + '/tests'
        img_for_test = os.path.join(project_root, img_for_test)

        mock_save_founded_certs.return_value = []

        # Llamar a la función a probar
        try:
            needed_certs = basic_process(img_for_test, data_type, url, origin_url, self.current_certs)
        except Exception as e:
            print(e)

        # Comparar el resultado con lo esperado
        self.assertEqual(res_certs, needed_certs)

    # @patch('app.data_process.save_founded_certs')
    # def test_basic_process_img_10(self, mock_save_founded_certs):
    #     res_certs = []
    #     data_type = 'img'
    #     img_for_test = 'dolar.jpg'
    #     needed_certs = []

        project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__))) + '/tests'
        img_for_test = os.path.join(project_root, img_for_test)
    #
    #     mock_save_founded_certs.return_value = []
    #
    #     # Llamar a la función a probar
    #     try:
    #         needed_certs = basic_process(img_for_test, data_type, url, origin_url, self.current_certs)
    #     except Exception as e:
    #         print(e)
    #
    #     # Comparar el resultado con lo esperado
    #     self.assertEqual(res_certs, needed_certs)

    @patch('app.data_process.save_founded_certs')
    def test_basic_process_img_11(self, mock_save_founded_certs):
        res_certs = []
        data_type = 'img'
        img_for_test = 'dolar.jpg'
        needed_certs = []

        project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__))) + '/tests'
        img_for_test = os.path.join(project_root, img_for_test)

        mock_save_founded_certs.return_value = []

        # Llamar a la función a probar
        try:
            needed_certs = basic_process(img_for_test, data_type, url, origin_url, self.current_certs)
        except Exception as e:
            print(e)

        # Comparar el resultado con lo esperado
        self.assertEqual(res_certs, needed_certs)

    @patch('app.data_process.save_founded_certs')
    def test_basic_process_img_12(self, mock_save_founded_certs):
        res_certs = []
        data_type = 'img'
        img_for_test = 'troncos.jpg'
        needed_certs = []

        project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__))) + '/tests'
        img_for_test = os.path.join(project_root, img_for_test)

        mock_save_founded_certs.return_value = []

        # Llamar a la función a probar
        try:
            needed_certs = basic_process(img_for_test, data_type, url, origin_url, self.current_certs)
        except Exception as e:
            print(e)

        # Comparar el resultado con lo esperado
        self.assertEqual(res_certs, needed_certs)

    @patch('app.data_process.save_founded_certs')
    def test_basic_process_img_13(self, mock_save_founded_certs):
        res_certs = []
        data_type = 'img'
        img_for_test = 'mar.jpg'
        needed_certs = []

        project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__))) + '/tests'
        img_for_test = os.path.join(project_root, img_for_test)

        mock_save_founded_certs.return_value = []

        # Llamar a la función a probar
        try:
            needed_certs = basic_process(img_for_test, data_type, url, origin_url, self.current_certs)
        except Exception as e:
            print(e)

        # Comparar el resultado con lo esperado
        self.assertEqual(res_certs, needed_certs)

    @patch('app.data_process.save_founded_certs')
    def test_basic_process_img_14(self, mock_save_founded_certs):
        res_certs = []
        data_type = 'img'
        img_for_test = 'hielo.jpg'
        needed_certs = []

        project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__))) + '/tests'
        img_for_test = os.path.join(project_root, img_for_test)

        mock_save_founded_certs.return_value = []

        # Llamar a la función a probar
        try:
            needed_certs = basic_process(img_for_test, data_type, url, origin_url, self.current_certs)
        except Exception as e:
            print(e)

        # Comparar el resultado con lo esperado
        self.assertEqual(res_certs, needed_certs)

    @patch('app.data_process.save_founded_certs')
    def test_basic_process_img_15(self, mock_save_founded_certs):
        res_certs = []
        data_type = 'img'
        img_for_test = 'tren.jpg'
        needed_certs = []

        project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__))) + '/tests'
        img_for_test = os.path.join(project_root, img_for_test)

        mock_save_founded_certs.return_value = []

        # Llamar a la función a probar
        try:
            needed_certs = basic_process(img_for_test, data_type, url, origin_url, self.current_certs)
        except Exception as e:
            print(e)

        # Comparar el resultado con lo esperado
        self.assertEqual(res_certs, needed_certs)

    @patch('app.data_process.save_founded_certs')
    def test_basic_process_img_16(self, mock_save_founded_certs):
        res_certs = []
        data_type = 'img'
        img_for_test = 'cueva.jpg'
        needed_certs = []

        project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__))) + '/tests'
        img_for_test = os.path.join(project_root, img_for_test)

        mock_save_founded_certs.return_value = []

        # Llamar a la función a probar
        try:
            needed_certs = basic_process(img_for_test, data_type, url, origin_url, self.current_certs)
        except Exception as e:
            print(e)

        # Comparar el resultado con lo esperado
        self.assertEqual(res_certs, needed_certs)

    @patch('app.data_process.save_founded_certs')
    def test_basic_process_img_17(self, mock_save_founded_certs):
        res_certs = []
        data_type = 'img'
        img_for_test = 'good.jpg'
        needed_certs = []

        project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__))) + '/tests'
        img_for_test = os.path.join(project_root, img_for_test)

        mock_save_founded_certs.return_value = []

        # Llamar a la función a probar
        try:
            needed_certs = basic_process(img_for_test, data_type, url, origin_url, self.current_certs)
        except Exception as e:
            print(e)

        # Comparar el resultado con lo esperado
        self.assertEqual(res_certs, needed_certs)

    @patch('app.data_process.save_founded_certs')
    def test_basic_process_img_18(self, mock_save_founded_certs):
        res_certs = ['S&P Global']
        data_type = 'img'
        img_for_test = 'spglobal.png'
        needed_certs = []

        project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__))) + '/tests'
        img_for_test = os.path.join(project_root, img_for_test)

        mock_save_founded_certs.return_value = []

        # Llamar a la función a probar
        try:
            needed_certs = basic_process(img_for_test, data_type, url, origin_url, self.current_certs)
        except Exception as e:
            print(e)

        # Comparar el resultado con lo esperado
        self.assertEqual(res_certs, needed_certs)

    @patch('app.data_process.save_founded_certs')
    def test_basic_process_img_19(self, mock_save_founded_certs):
        res_certs = []
        data_type = 'img'
        img_for_test = 'false_1.png'
        needed_certs = []

        project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__))) + '/tests'
        img_for_test = os.path.join(project_root, img_for_test)

        mock_save_founded_certs.return_value = []

        # Llamar a la función a probar
        try:
            needed_certs = basic_process(img_for_test, data_type, url, origin_url, self.current_certs)
        except Exception as e:
            print(e)

        # Comparar el resultado con lo esperado
        self.assertEqual(res_certs, needed_certs)

    @patch('app.data_process.save_founded_certs')
    def test_basic_process_img_20(self, mock_save_founded_certs):
        res_certs = []
        data_type = 'img'
        img_for_test = 'nave.jpg'
        needed_certs = []

        project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__))) + '/tests'
        img_for_test = os.path.join(project_root, img_for_test)

        mock_save_founded_certs.return_value = []

        # Llamar a la función a probar
        try:
            needed_certs = basic_process(img_for_test, data_type, url, origin_url, self.current_certs)
        except Exception as e:
            print(e)

        # Comparar el resultado con lo esperado
        self.assertEqual(res_certs, needed_certs)

    @patch('app.data_process.save_founded_certs')
    def test_basic_process_img_21(self, mock_save_founded_certs):
        res_certs = []
        data_type = 'img'
        img_for_test = 'hombre.png'
        needed_certs = []

        project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__))) + '/tests'
        img_for_test = os.path.join(project_root, img_for_test)

        mock_save_founded_certs.return_value = []

        # Llamar a la función a probar
        try:
            needed_certs = basic_process(img_for_test, data_type, url, origin_url, self.current_certs)
        except Exception as e:
            print(e)

        # Comparar el resultado con lo esperado
        self.assertEqual(res_certs, needed_certs)

    # @patch('app.data_process.save_founded_certs')
    # def test_basic_process_img_22(self, mock_save_founded_certs):
    #     res_certs = ['CDP']
    #     data_type = 'img'
    #     img_for_test = 'cdp.png'
    #
    #     mock_save_founded_certs.return_value = []
    #
    #     # Llamar a la función a probar
    #     try:
    #         needed_certs = basic_process(img_for_test, data_type, url, origin_url, self.current_certs)
    #     except Exception as e:
    #         print(e)
    #
    #     # Comparar el resultado con lo esperado
    #     self.assertEqual(res_certs, needed_certs)

    @patch('app.data_process.save_founded_certs')
    def test_basic_process_img_23(self, mock_save_founded_certs):
        res_certs = ['Ecovadis Platinum']
        data_type = 'img'
        img_for_test = 'ecovadis_plat.png'
        needed_certs = []

        project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__))) + '/tests'
        img_for_test = os.path.join(project_root, img_for_test)

        mock_save_founded_certs.return_value = []

        # Llamar a la función a probar
        try:
            needed_certs = basic_process(img_for_test, data_type, url, origin_url, self.current_certs)
        except Exception as e:
            print(e)

        # Comparar el resultado con lo esperado
        self.assertEqual(res_certs, needed_certs)

    @patch('app.data_process.save_founded_certs')
    def test_basic_process_img_24(self, mock_save_founded_certs):
        res_certs = []
        data_type = 'img'
        img_for_test = 'nave2.jpg'
        needed_certs = []

        project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__))) + '/tests'
        img_for_test = os.path.join(project_root, img_for_test)

        mock_save_founded_certs.return_value = []

        # Llamar a la función a probar
        try:
            needed_certs = basic_process(img_for_test, data_type, url, origin_url, self.current_certs)
        except Exception as e:
            print(e)

        # Comparar el resultado con lo esperado
        self.assertEqual(res_certs, needed_certs)

    @patch('app.data_process.save_founded_certs')
    def test_basic_process_pdf1(self, mock_save_founded_certs):
        res_certs = []
        data_type = 'doc'
        pdf_for_test = 'false_1.pdf'
        needed_certs = []

        project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__))) + '/tests'
        pdf_for_test = os.path.join(project_root, pdf_for_test)

        mock_save_founded_certs.return_value = []

        # Llamar a la función a probar
        try:
            needed_certs = basic_process(pdf_for_test, data_type, url, origin_url, self.current_certs)
        except Exception as e:
            print(e)

        # Comparar el resultado con lo esperado
        self.assertEqual(res_certs, needed_certs)

    # @patch('app.data_process.save_founded_certs')
    # def test_basic_process_pdf2(self, mock_save_founded_certs):
    #     res_certs = ['ISO 9001']
    #     data_type = 'doc'
    #     pdf_for_test = 'true_1.pdf'
    #
    #     mock_save_founded_certs.return_value = []
    #
    #     # Llamar a la función a probar
    #     try:
    #         needed_certs = basic_process(pdf_for_test, data_type, url, origin_url, self.current_certs)
    #     except Exception as e:
    #         print(e)
    #
    #     # Comparar el resultado con lo esperado
    #     self.assertEqual(res_certs, needed_certs)
    #
    # @patch('app.data_process.save_founded_certs')
    # def test_basic_process_pdf3(self, mock_save_founded_certs):
    #     res_certs = ['ISO 9001']
    #     data_type = 'doc'
    #     pdf_for_test = 'true_2.pdf'
    #
    #     mock_save_founded_certs.return_value = []
    #
    #     # Llamar a la función a probar
    #     try:
    #         needed_certs = basic_process(pdf_for_test, data_type, url, origin_url, self.current_certs)
    #     except Exception as e:
    #         print(e)
    #
    #     # Comparar el resultado con lo esperado
    #     self.assertEqual(res_certs, needed_certs)


if __name__ == '__main__':
    unittest.main()
