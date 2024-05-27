import io
import json
import os
import sys
import time
import pika
import logging
import requests
import fitz
import cv2

from PIL import Image, ImageFile, UnidentifiedImageError, ImageEnhance
from langdetect import detect, detect_langs
import spacy

import pytesseract

from fuzzywuzzy import fuzz, process
from openpyxl import load_workbook

API_URL = "http://localhost:8000/api"

host = 'rat.rmq2.cloudamqp.com'
username = 'qojzruiu'
password = 'apgHrzsZgzGr1SBXE9DwJ3WvhuiS0Raz'

credentials = pika.PlainCredentials(username, password)
# Conexión al servidor RabbitMQ en localhost
connection_params = pika.ConnectionParameters(host, 5672, 'qojzruiu', credentials)

logging.basicConfig(stream=sys.stdout, level=logging.INFO)

logging.getLogger("pika").setLevel(logging.WARNING)

pytesseract.pytesseract.tesseract_cmd = r'C:\Users\T031105\AppData\Local\Programs\Tesseract-OCR\tesseract.exe'

# ImageFile.LOAD_TRUNCATED_IMAGES = True

certs_to_find = []

whitelist_chars = " &'(),-.0123456789:ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz ­ÁÂÍÏÑÓáâíïñó"


def get_circular_text_from_img(img_src):
    found_text = ''

    # Read image
    img = cv2.imread(img_src)

    # cv2.imwrite('img2.jpg', img)

    # Convert to grayscale, and binarize, especially for removing JPG artifacts
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    gray = cv2.threshold(gray, 128, 255, cv2.THRESH_BINARY_INV)[1]

    # Crop center part of image to simplify following contour detection
    h, w = gray.shape
    l = (w - h) // 2
    gray = gray[:, l:l + h]

    # Find (nested) contours (cf. cv2.RETR_TREE) w.r.t. the OpenCV version
    cnts = cv2.findContours(gray, cv2.RETR_TREE, cv2.CHAIN_APPROX_NONE)
    cnts = cnts[0] if len(cnts) == 2 else cnts[1]

    # Filter and sort contours on area
    cnts = [cnt for cnt in cnts if cv2.contourArea(cnt) > 10000]
    cnts = sorted(cnts, key=cv2.contourArea)

    # Remove inner text by painting over using found contours
    # Contour index 1 = outer edge of inner circle
    gray = cv2.drawContours(gray, cnts, 1, 0, cv2.FILLED)

    # If specifically needed, also remove text in the original image
    # Contour index 0 = inner edge of inner circle (to keep inner circle itself)
    img[:, l:l + h] = cv2.drawContours(img[:, l:l + h], cnts, 0, (255, 255, 255),
                                       cv2.FILLED)

    # Rotate image before remapping to polar coordinate space to maintain
    # circular text en bloc after remapping
    gray = cv2.rotate(gray, cv2.ROTATE_90_COUNTERCLOCKWISE)

    center = (int(h // 2.45), int(h // 2))

    # cv2.circle(gray, center, 5, (0, 0, 0), -1)

    # cv2.imwrite('img1.jpg', gray)

    max_radius = int(h * 0.5)

    # Actual remapping to polar coordinate space
    gray = cv2.warpPolar(gray, (-1, -1), center, max_radius,
                         cv2.INTER_CUBIC + cv2.WARP_POLAR_LINEAR)
    # cv2.imwrite('img2.jpg', gray)
    # Rotate result for OCR
    gray = cv2.rotate(gray, cv2.ROTATE_90_COUNTERCLOCKWISE)

    # cv2.imwrite('img3.jpg', gray)

    # Actual OCR, limiting to capital letters only
    config = f'--psm 6 -c tessedit_char_whitelist="{whitelist_chars}"'
    text = pytesseract.image_to_string(gray, config=config)

    found_text += text.replace('\n', '').replace('\f', '')

    return found_text


def extract_text(image_src):
    text = ''

    try:
        image = Image.open(image_src)

        try:
            text += get_circular_text_from_img(image_src)
        except Exception as e:
            print(e)

        # Uso de Tesseract con configuraciones mejoradas
        custom_config = f'--oem 3 --psm 6 -c tessedit_char_whitelist="{whitelist_chars}"'
        text += pytesseract.image_to_string(image, lang='eng', config=custom_config)

        logging.debug(text)

        return text.strip()

    except Exception as e:
        logging.debug(str(e))
        logging.debug(f"Ocurrió un error inesperado: {str(e)}")


def extract_text_from_pdf(src, certs):
    found_certs = []

    try:
        # Abrir el archivo PDF
        pdf_document = fitz.open(src)
    except Exception as e:
        logging.debug(f"Error al abrir el archivo PDF: {str(e)}")
        return

    for page_num in range(pdf_document.page_count):
        print(f'Procesando pagina '+str(page_num))
        try:
            page = pdf_document.load_page(page_num)

            # Extraer y mostrar el texto
            text = page.get_text("text")

            found_certs += find_certs_in_text(text, certs, True, 90)

            # Extraer y guardar imágenes
            images = page.get_images(full=True)
            for img_index, img in enumerate(images):
                try:
                    xref = img[0]
                    base_image = pdf_document.extract_image(xref)
                    if base_image:
                        image_bytes = base_image["image"]
                        image_ext = base_image["ext"]

                        # Verificar si los bytes extraídos son una imagen válida
                        image = Image.open(io.BytesIO(image_bytes))

                        image_text = pytesseract.image_to_string(image, lang='eng')

                        logging.debug('TEXTO ENCONTRADO EN PDF ')
                        logging.debug(image_text)

                        found_certs += find_certs_in_text(image_text, certs, False, 85)

                except UnidentifiedImageError:
                    logging.debug(
                        f"No se pudo identificar la imagen en la página {page_num + 1}, imagen {img_index + 1}")
                except IOError as e:
                    logging.debug(
                        f"Error al guardar la imagen en la página {page_num + 1}, imagen {img_index + 1}: {str(e)}")
                except Exception as e:
                    logging.debug(
                        f"Ocurrió un error inesperado al procesar la imagen en la página {page_num + 1}, imagen {img_index + 1}: {str(e)}")

        except Exception as e:
            logging.debug(f"Ocurrió un error al procesar la página {page_num + 1}: {str(e)}")

    return found_certs


def basic_process(data_src, data_type, url, origin_url, certs):
    found_certs = set()

    text = ''

    nlp = False

    umbral = 82

    print('Certificados encontrados: '+ data_src)
    print(list(found_certs))

    match data_type:
        case 'txt':
            nlp = True
            umbral = 90
            with open(data_src, 'r', encoding='utf-8') as file:
                text = file.read()

        case 'img':
            umbral = 85
            text = extract_text(data_src)

        case 'doca':
            nlp = True
            _, file_extension = os.path.splitext(data_src)

            if file_extension == '.pdf':
                ex_certs = extract_text_from_pdf(data_src, certs)

                if ex_certs is not None:
                    for ex_cert in ex_certs:
                        found_certs.add(ex_cert)


                # if text_from_image:
                #     certs_into_images = find_certs_in_text(text_from_image, certs, False)
                #
                #     for cert in certs_into_images:
                #         found_certs.append(cert)

        case _:
            logging.debug(f'Unsupported data type: {data_type}')



    found_and_filtered_certs = []

    if text:
        ex_certs = find_certs_in_text(text, certs, nlp, umbral)

        if ex_certs is not None:
            for ex_cert in ex_certs:
                found_certs.add(ex_cert)


    if found_certs:
        try:
            found_and_filtered_certs = save_founded_certs(list(found_certs), data_src, data_type, origin_url, url)
        except Exception as e:
            print(e)

    return list(found_certs)


def save_founded_certs(found_certs, data_src, data_type, origin_url, url):
    print(found_certs)

    certs = []

    response_url = requests.get(f"{API_URL}/urls/{origin_url}")

    if response_url.status_code != 200:
        logging.error(f"Error al obtener la company")
        return

    for cert in found_certs:
        logging.debug(f'Found certificate: {cert} ')

        response_cert = requests.get(f"{API_URL}/certificates?name={cert}")

        data_cert = response_cert.json()

        if data_cert:
            data_cert = data_cert[0]
        else:
            data_cert = requests.post(f'{API_URL}/certificates',
                                      json={'name': cert}).json()

        data_type = data_type.upper()

        response = requests.post(f'{API_URL}/resources',
                                 json={'type': data_type, 'url_id': origin_url, 'path_file': data_src,
                                       'full_url': url, 'certificate_id': data_cert.get('id')})

        certs.append(cert)

    return certs


def callback(ch, method, properties, body):
    logging.info(f" [x] Received {body}")

    message = json.loads(body)

    data_src = message.get('data_src')
    data_type = message.get('type')
    url = message.get('url')
    origin_url = message.get('origin_url')

    ch.basic_ack(delivery_tag=method.delivery_tag)

    start_time = time.time()

    basic_process(data_src, data_type, url, origin_url, certs_to_find)

    end_time = time.time()
    elapsed_time = end_time - start_time
    print(f"Tiempo transcurrido en basic_process: {elapsed_time:.2f} segundos")




def consume_messages(queue='default'):
    while True:
        try:
            # Conectar al servidor RabbitMQ
            connection = pika.BlockingConnection(connection_params)
            channel = connection.channel()

            # Declarar una cola llamada 'hello'
            channel.queue_declare(queue=queue)

            # Suscribirse a la cola
            channel.basic_consume(queue=queue,
                                  on_message_callback=callback,
                                  auto_ack=False)

            logging.info(' [*] Waiting for messages. To exit press CTRL+C')
            channel.start_consuming()

        except pika.exceptions.AMQPConnectionError:
            print("Error de conexión, reintentando en 5 segundos...")
            time.sleep(5)


def unique_characters(strings):
    unique_chars = set()  # Utilizamos un conjunto para evitar caracteres repetidos
    for string in strings:
        for char in string:
            unique_chars.add(char)
            if char.islower():
                unique_chars.add(char.upper())
            elif char.isupper():
                unique_chars.add(char.lower())

    unique_chars.add(' ')

    sorted_chars = sorted(unique_chars)

    return ''.join(sorted_chars)


def find_certs_in_text(text, certs, use_nlp, umbral=82):
    print('[*] Finding certs... '+str(use_nlp))

    found_certs = []

    print(text)

    if text:
        if use_nlp:
            lang = 'en'
            try:
                lang = detect(text)
            except Exception as e:
                logging.debug(e)

            nlp = None

            logging.debug(text)
            logging.debug(lang)

            if lang == 'es':
                nlp = spacy.load("es_core_news_sm")
            else:
                nlp = spacy.load("en_core_web_sm")

            nlp.max_length = 20000000

            # Procesar el texto
            doc = nlp(text)

            found_certs = [ent.text for ent in doc.ents if ent.text in certs]

            # print(found_certs)
        else:
            text = text.replace('\n', ' ').replace('\r', '').replace(' ', '')
            # print(text)
            best_cert_score = 0
            best_cert = None
            for cert in certs:
                matches = process.extract(cert.lower().replace(' ', ''), [word.lower() for word in text.split() if len(word) > 4],
                                          limit=1,
                                          scorer=fuzz.partial_ratio)
                # matches = process.extract(cert, text, limit=2)
                for match in matches:
                    # print(cert + ' ' + str(match[1]))
                    if match[1] >= umbral and match[
                        1] > best_cert_score:  # 75 es un umbral de similitud, puedes ajustarlo según sea necesario
                        best_cert_score = match[1]
                        best_cert = cert

            if best_cert is not None:
                found_certs.append(best_cert)

        print(found_certs)

    return found_certs



if __name__ == "__main__":
    logging.info('Procesando datos en la cola')

    wb = load_workbook('Glosario_Certificaciones.xlsx')
    sheet = wb.active
    # Leer todas las filas de la columna B y convertirlas en un array de certificados
    for row in sheet.iter_rows(min_row=2, min_col=2, max_col=2, values_only=True):
        cell_value = row[0]
        if cell_value is not None:
            certs_to_find.append(cell_value)

    # " FECIGZÓacráHyNlT A:­Â(d)o&.8ï'bYK2WzvgXÁâ3ih,SJuÏMO-ejk6L7t9mRqóÍ4fDñÑBín5pPwsUVQ10x"
    whitelist_chars = unique_characters(certs_to_find)

    consume_messages('data_process')
